import spacy
from SuffixTrie import maxMatch
import en_core_web_sm
import string
from openpyxl import load_workbook

nlp = en_core_web_sm.load()


def toLemmas(text):
    doc = nlp(text)
    return " ".join([token.lemma_ for token in doc])


def extract_unlemmatized_nouns(doc, strie):
    ul_nouns = []
    for chunk in doc.noun_chunks:
        if chunk.root.dep_ == "nsubj" and chunk.root.head.lemma_ == "be" and doc[chunk.root.head.i + 1].pos_ == "DET":
            result_span = get_term_in_noun_chunk(strie, chunk, doc)
            phrase = doc.char_span(result_span[0], result_span[1]).text.lower()
            ul_nouns.append(phrase)
    return ul_nouns


def extract_definitions(filename):
    wb = load_workbook(filename=filename)
    ws = wb.active
    definitions = []
    for row in range(sum(1 for _ in ws.rows) - 2):
        cell_dim = 'B' + str(row + 3)
        definitions.append(ws[cell_dim].value)
    return definitions


def defined_nouns_from_av2(filename, phraseLinks):
    defined_nouns = set()
    acronymLinks = {v: k for k, v in phraseLinks.items()}
    definitions = extract_definitions(filename)
    for definition in definitions:
        doc = nlp(definition)
        for chunk in doc.noun_chunks:
            regToken = chunk.root.text.lower().translate(str.maketrans('', '', string.punctuation))
            if regToken in phraseLinks:
                defined_nouns.add(phraseLinks[regToken])
        lemmatizedPhrase = toLemmas(definition.lower())
        defined_nouns.add(lemmatizedPhrase)
        if lemmatizedPhrase in acronymLinks:
          defined_nouns.add(acronymLinks[lemmatizedPhrase])
    return defined_nouns


def extract_defined_nouns(doc, strie, acronymLinks, phraseLinks):
    defined_nouns = set()
    for chunk in doc.noun_chunks:
        if chunk.root.dep_ == "nsubj" and chunk.root.head.lemma_ == "be" and doc[chunk.root.head.i + 1].pos_ == "DET":
            result_span = get_term_in_noun_chunk(strie, chunk, doc)
            phrase = doc.char_span(result_span[0], result_span[1]).text.lower()
            lemmatizedPhrase = toLemmas(phrase)
            defined_nouns.add(lemmatizedPhrase)
            if lemmatizedPhrase in acronymLinks:
                defined_nouns.add(acronymLinks[lemmatizedPhrase])
            regToken = chunk.root.text.lower().translate(str.maketrans('', '', string.punctuation))
            if regToken in phraseLinks:
                defined_nouns.add(phraseLinks[regToken])
    return defined_nouns


def extract_defined_nouns_from_all(docs, strie, phraseLinks):
    defined_nouns = set()
    acronymLinks = {v: k for k, v in phraseLinks.items()}
    for doc in docs:
        defined_nouns = defined_nouns.union(extract_defined_nouns(doc, strie, acronymLinks, phraseLinks))
    return defined_nouns


def get_term_in_noun_chunk(trie, chunk, doc):
    maxMatchReturn = maxMatch(trie, chunk)
    if maxMatchReturn is None:
        end = doc[chunk.end - 1].idx + len(doc[chunk.end - 1].text)
        if doc[chunk.start].pos_ == "DET":
            start = doc[chunk.start + 1].idx
            return (start, end)
        else:
            start = doc[chunk.start].idx
            return (start, end)
    return maxMatchReturn
